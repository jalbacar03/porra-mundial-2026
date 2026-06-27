"""
Backup completo de predicciones -> XLSX (3 hojas).
Participantes en filas, predicciones en columnas. Incluye Bot365.
Lee DATABASE_URL de .env.backup. Uso: .venv-backup/bin/python scripts/export_backup.py
"""
import os, re, datetime
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BOT365 = 'b0365b03-65b0-365b-0365-b0365b036500'

# --- Cargar DATABASE_URL de .env.backup ---
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
env_path = os.path.join(ROOT, '.env.backup')
url = None
with open(env_path) as f:
    for line in f:
        line = line.strip()
        if line.startswith('DATABASE_URL='):
            url = line.split('=', 1)[1].strip().strip('"').strip("'")
if not url:
    raise SystemExit('No encuentro DATABASE_URL en .env.backup')

conn = psycopg2.connect(url)
cur = conn.cursor()

# --- Catálogos ---
cur.execute("SELECT id, name FROM teams")
team_name = {r[0]: r[1] for r in cur.fetchall()}

# Participantes: admitidos (has_paid) + Bot365. Bot365 al final.
cur.execute("""
  SELECT id, full_name, (id = %s) AS is_bot
  FROM profiles
  WHERE has_paid = true OR id = %s
  ORDER BY (id = %s), lower(full_name)
""", (BOT365, BOT365, BOT365))
participants = cur.fetchall()  # (id, full_name, is_bot)

# --- Estilos ---
HEAD_FILL = PatternFill('solid', fgColor='1A1D26')
HEAD_FONT = Font(color='FFCC00', bold=True, size=10)
NAME_FONT = Font(bold=True)
BOT_FILL = PatternFill('solid', fgColor='FFF7D6')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = CENTER
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}1"
    ws.column_dimensions['A'].width = 26
    ws.row_dimensions[1].height = 42

wb = Workbook()

# ============ HOJA 1: GRUPOS ============
ws = wb.active
ws.title = 'Grupos'
cur.execute("""
  SELECT m.id, m.match_number, m.group_name, th.name, ta.name
  FROM matches m
  JOIN teams th ON th.id = m.home_team_id
  JOIN teams ta ON ta.id = m.away_team_id
  WHERE m.stage = 'group'
  ORDER BY m.match_number
""")
gmatches = cur.fetchall()  # (id, num, group, home, away)
gcols = [r[0] for r in gmatches]
ws.cell(1, 1, 'Participante')
for j, (mid, num, grp, h, a) in enumerate(gmatches, start=2):
    ws.cell(1, j, f"G{grp} · {h}-{a}")

# Todas las predicciones de grupo: {(user, match): "h-a"}
cur.execute("""
  SELECT pr.user_id, pr.match_id, pr.predicted_home, pr.predicted_away
  FROM predictions pr
  JOIN matches m ON m.id = pr.match_id
  WHERE m.stage = 'group'
""")
gpred = {}
for uid, mid, h, a in cur.fetchall():
    if h is not None and a is not None:
        gpred[(uid, mid)] = f"{h}-{a}"

for i, (uid, name, is_bot) in enumerate(participants, start=2):
    ws.cell(i, 1, name).font = NAME_FONT
    for j, mid in enumerate(gcols, start=2):
        v = gpred.get((uid, mid), '')
        c = ws.cell(i, j, v)
        c.alignment = CENTER
        if is_bot:
            c.fill = BOT_FILL
    if is_bot:
        ws.cell(i, 1).fill = BOT_FILL
for j in range(2, len(gcols) + 2):
    ws.column_dimensions[get_column_letter(j)].width = 12
style_header(ws, len(gcols) + 1)

# ============ HOJA 2: CUADRO ============
ws2 = wb.create_sheet('Cuadro')
round_order = {'r32': 0, 'r16': 1, 'qf': 2, 'sf': 3, 'final': 4}
round_label = {'r32': '16avos', 'r16': 'Octavos', 'qf': 'Cuartos', 'sf': 'Semis', 'final': 'Final'}
cur.execute("SELECT DISTINCT round, match_number FROM bracket_picks")
bslots = sorted({(r, n) for r, n in cur.fetchall()},
                key=lambda x: (round_order.get(x[0], 9), x[1]))
ws2.cell(1, 1, 'Participante')
for j, (rnd, num) in enumerate(bslots, start=2):
    ws2.cell(1, j, f"{round_label.get(rnd, rnd)} #{num}")

cur.execute("SELECT user_id, round, match_number, predicted_winner_id FROM bracket_picks")
bpick = {}
for uid, rnd, num, wid in cur.fetchall():
    bpick[(uid, rnd, num)] = team_name.get(wid, '') if wid else ''

for i, (uid, name, is_bot) in enumerate(participants, start=2):
    ws2.cell(i, 1, name).font = NAME_FONT
    for j, (rnd, num) in enumerate(bslots, start=2):
        c = ws2.cell(i, j, bpick.get((uid, rnd, num), ''))
        c.alignment = CENTER
        if is_bot:
            c.fill = BOT_FILL
    if is_bot:
        ws2.cell(i, 1).fill = BOT_FILL
for j in range(2, len(bslots) + 2):
    ws2.column_dimensions[get_column_letter(j)].width = 14
style_header(ws2, len(bslots) + 1)

# ============ HOJA 3: ESPECIALES ============
ws3 = wb.create_sheet('Especiales')
cur.execute("SELECT id, slug, name, input_type FROM pre_tournament_bets WHERE is_active = true ORDER BY sort_order, id")
bets = cur.fetchall()  # (id, slug, name, input_type)
ws3.cell(1, 1, 'Participante')
for j, (bid, slug, nm, itype) in enumerate(bets, start=2):
    ws3.cell(1, j, nm)

cur.execute("SELECT user_id, bet_id, value FROM pre_tournament_entries")
def readable(value, itype):
    if value is None:
        return ''
    v = value if isinstance(value, dict) else {}
    if 'team_id' in v and v['team_id'] is not None:
        return team_name.get(v['team_id'], str(v['team_id']))
    if 'player_name' in v and v['player_name']:
        return v['player_name']
    if 'answer' in v and v['answer'] is not None:
        a = str(v['answer']).lower()
        return {'yes': 'Sí', 'no': 'No'}.get(a, str(v['answer']))
    return str(v) if v else ''
epick = {}
for uid, bid, value in cur.fetchall():
    itype = next((b[3] for b in bets if b[0] == bid), None)
    epick[(uid, bid)] = readable(value, itype)

for i, (uid, name, is_bot) in enumerate(participants, start=2):
    ws3.cell(i, 1, name).font = NAME_FONT
    for j, (bid, slug, nm, itype) in enumerate(bets, start=2):
        c = ws3.cell(i, j, epick.get((uid, bid), ''))
        c.alignment = CENTER
        if is_bot:
            c.fill = BOT_FILL
    if is_bot:
        ws3.cell(i, 1).fill = BOT_FILL
for j in range(2, len(bets) + 2):
    ws3.column_dimensions[get_column_letter(j)].width = 22
style_header(ws3, len(bets) + 1)

# --- Guardar ---
os.makedirs(os.path.join(ROOT, 'exports'), exist_ok=True)
today = datetime.date.today().isoformat()
out = os.path.join(ROOT, 'exports', f'porra_backup_{today}.xlsx')
wb.save(out)

# --- Resumen / verificación ---
n_bot = sum(1 for p in participants if p[2])
print(f"OK -> {out}")
print(f"Participantes: {len(participants)} (incluido Bot365: {'sí' if n_bot else 'NO'})")
print(f"Grupos: {len(gcols)} columnas · {len(gpred)} predicciones totales")
print(f"Cuadro: {len(bslots)} columnas · {len(bpick)} picks totales")
print(f"Especiales: {len(bets)} columnas · {len(epick)} respuestas totales")
cur.close(); conn.close()
